import unittest
from unittest.mock import patch
from intrinio_sdk.rest import ApiException
from exception.exceptions import ReportError, ValidationError, FileSystemError
from reporting.workbook_report import WorkbookReport
from reporting.jimmy_report_worksheet import JimmyReportWorksheet
from valuation_models.jimmy_model import JimmyValuationModel
from openpyxl import Workbook
from openpyxl import load_workbook
import os.path
import os

report_output_override = "../test/spreadsheet/"
test_out_name = "out.xlsx"

class TestWorkbookReport(unittest.TestCase):

    def test_report_no_sheets(self):

        report = WorkbookReport(report_output_override)

        #do not add any worksheets
        with self.assertRaises(ValidationError):
            report.generate_report("bad-report.xlsx")

    def test_report_no_reportName(self):
        report = WorkbookReport(report_output_override)

        report.add_worksheet(JimmyReportWorksheet(), "Test Spreadsheet", JimmyValuationModel('appl', 2019))

        with self.assertRaises(ValidationError):
            report.generate_report(None)

        with self.assertRaises(ValidationError):
            report.generate_report("")

    def test_save_invalid_output_path(self):
        with self.assertRaises(FileSystemError):
            WorkbookReport("")

                

    def test_valid_report(self):
        '''
            This test is a bit more complex

            1) create a valid report in the test directory
            2) test that the file exists
            3) test that it can be read and that it contains a single sheet
        '''
        report = WorkbookReport(report_output_override)

        test_report = JimmyReportWorksheet()
        test_model = JimmyValuationModel('appl', 2019)
        report.add_worksheet(test_report, 'test_sheet', test_model)

        wb = Workbook()
        ws = wb.create_sheet('test_sheet')

        with patch.object(test_report, 'prepare_worksheet', return_value=ws), \
             patch.object(test_model, 'calculate_dcf_price', return_value=100):
            
            report.generate_report(test_out_name)

            try:
                # verify that the report exists and has 1 sheet
                self.assertTrue(os.path.isfile(report_output_override + test_out_name))

                wb = load_workbook(filename='%s%s' %
                                (report_output_override, test_out_name))
                wb.close()

                self.assertEqual(len(wb.worksheets), 1)
            finally:
                os.remove(report_output_override + test_out_name)


            

    



        

    

