"""Author: Mark Hanegraaff -- 2019
"""
import os
import os.path
import logging
from exception.exceptions import ValidationError, ReportError
from support import util
from copy import copy

from openpyxl import Workbook

# can be overridden for testing purposed

log = logging.getLogger()

class WorkbookReport():
    """ 
        A class representing a spreadsheet workbook. Each workbook may contain
        one or more worksheets (spreadsheets) represnting the the output of
        some calculation

        Attributes: 
            output_path : str
                Path of the output report

            worksheet_list : str
                List of worksheets associated with this report.
                each worksheet is a tuple of a report object and
                a model that can perform the calculation to be reported

            price_dict : str
                A dictionary of intrinsic prices that are produced by
                the model associated with each worksheet

                price_dict[]

    """

    def __init__(self, output_path_override):
        """
            Creates the report directory if not already present

            Parameters
            ------------
            None

            Raises
            ------------
            A FileSystemError error in case the output path could not be created
            
            Returns
            ------------
            None
        """
        self.output_path = "./reports/"

        if output_path_override is not None:
            self.output_path = output_path_override
    
        util.create_dir(self.output_path)
        self.worksheet_list = []

        self.price_dict = {}


    def add_worksheet(self, report_worksheet : object, worksheet_title : str, dcf_model : object):
        """
            Adds a worksheet to the report

            Parameters
            ------------
            report_worksheet : object
                The report worksheet for the supplied model

            worksheet_tile : str
                The user friendly title used to name the worksheet

            dcf_model : object
                The DCF model ojbect that will perform the calculation

            Returns
            ------------
            None

        """
        self.worksheet_list.append((report_worksheet, worksheet_title, dcf_model))

    def generate_report(self, report_filename: str):
        """
            Generates a report and all associated worksheets

            Parameters
            ----------
            report_filename : str
                Name of the output report

            Raises
            ----------
            ValidationError
                In case the report could not be generate because it was not properly
                initialized
            ReportError
                In case of any errors preparing or generating the report.

            Returns
            -------
            None

        """

        wb = Workbook()
        wb.remove(wb.active)

        #
        # Assemble the report
        #

        if len(self.worksheet_list) == 0:
            raise ValidationError("No worksheets were supplied to the report", None)

        if report_filename is None or report_filename == "":
            raise ValidationError("No report filename was supplied", None)
        
        for (report_worksheet, worksheet_tile, dcf_model) in self.worksheet_list:
            self.price_dict[worksheet_tile] = dcf_model.calculate_dcf_price()
            
            source_worksheet = report_worksheet.create_worksheet(dcf_model.get_itermediate_results())
            target_worksheet = wb.create_sheet(worksheet_tile)
            
            self.__copy_worksheet__(source_worksheet, target_worksheet)
        
        output_report_name = '%s%s' % (self.output_path, report_filename)

        try:
            wb.save(output_report_name)
        except Exception as e:
            raise ReportError("Error saving report", e)


    def __copy_worksheet__(self, source_worksheet : object, dest_work_sheet : object):
        """
            Copies a worksheet cell by cell. Used to copy a worksheet from one
            spreadsheet into another

            Parameters
            ------------
            source_worksheet : object
                The source worksheet. Typically from a template spreadhseet

            dest_work_sheet : object
                The destination worksheet. The one that is being assembled

            Returns
            ------------
            None

        """
        
        # copy sheet values and styles
        for col_cells in source_worksheet.iter_cols():
            for cell in col_cells:
                dest_cell = dest_work_sheet.cell(row=cell.row, column=cell.col_idx, value=cell.value)
                if cell.has_style:
                    dest_cell.font = copy(cell.font)
                    dest_cell.border = copy(cell.border)
                    dest_cell.fill = copy(cell.fill)
                    dest_cell.number_format = copy(cell.number_format)
                    dest_cell.protection = copy(cell.protection)
                    dest_cell.alignment = copy(cell.alignment)

        # copy the column dimensions
        for index, dimensions in source_worksheet.column_dimensions.items():
            dest_work_sheet.column_dimensions[index].width = dimensions.width

        # copy the row dimensions
        for index, dimensions in source_worksheet.row_dimensions.items():
            dest_work_sheet.row_dimensions[index].height = dimensions.height
